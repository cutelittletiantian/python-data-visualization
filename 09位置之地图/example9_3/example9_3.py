import os
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pyecharts.charts import Map, Timeline
import pyecharts.options as opts

# 资源文件统一所在目录
resourcesPath = "./resources"
# 降雨量数据工作簿名
gpdBookName = "2010-2017年各省份历年GDP.xlsx"
# 打开文件
gpdBook = openpyxl.load_workbook(filename=os.path.join(resourcesPath, gpdBookName))  # type: Workbook


def get_data(year: str) -> tuple:
    """
    指定工作表名year，返回绘图所需的数据

    :param year: 年份，是字符串

    :return: 绘图所需数据，格式(年份——数值类型的年份, [(省份1, GDP取值1), (省份2, GDP取值2), ...])
    """
    # 选中工作表
    year_sheet = gpdBook[year]  # type: Worksheet

    # 取出全国当前的各省GDP数据
    gdp_list = [row_item for row_item in year_sheet.iter_rows(min_row=2, values_only=True)]
    # 清洗数据，去掉含空项
    gdp_list = [item for item in gdp_list if None not in item]

    return (
        eval(str(year)),
        gdp_list
    )


# 扫描各个工作表，绘制轮播图
timeline = Timeline()
for yearSheetName in gpdBook.sheetnames:
    # 取出绘图所需的数据
    year, dataPair = get_data(year=yearSheetName)

    # 绘制单组地图需求：
    # 1. 地图类型设置为中国地图；
    # 2. 视觉映射配置项最大映射值设置为3000；
    # 3. 标题设置为f"{data[0]}年各省GDP（单位：10亿人民币）"；
    # 4. 轮播图时间节点设置为“xxxx年”
    # 5. 隐藏图例
    timeline.add(
        time_point=f"{year}年",
        chart=(
            Map()
            .set_global_opts(
                title_opts=opts.TitleOpts(title=f"{year}年各省GDP（单位：10亿人民币）"),
                visualmap_opts=opts.VisualMapOpts(max_=3000)
            )
            .add(
                series_name=f"{year}年",
                maptype="china",
                data_pair=dataPair
            )
        )
    )

# 保存路径
resultPath = "./result"
if not os.path.exists(path=resultPath):
    os.mkdir(path=resultPath)
# 保存文件名
resultFileName = "gdp_map_timeline.html"
# 执行保存
timeline.render(path=os.path.join(resultPath, resultFileName))
