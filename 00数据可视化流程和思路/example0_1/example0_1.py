# 使用import导入openpyxl模块
import openpyxl
# 使用from...import从pyecharts.charts导入Sankey
from pyecharts.charts import Sankey
# 使用from...import从pyecharts导入options，简写为opts
from pyecharts import options as opts
# 使用from...import从pyecharts.charts导入Page
from pyecharts.charts import Page

# 使用openpyxl.load_workbook(path)读取文件，赋值给wb
wb = openpyxl.load_workbook("./resources/sankey.xlsx")
# 使用中括号读取工作表对照组和实验组，赋值给sheet_A,sheet_B
sheet_A = wb["对照组"]
sheet_B = wb["实验组"]


# 定义函数gen_nodesandlinks传入参数sheet
def gen_nodesandlinks(sheet):
    # 定义标签数据列表，赋值给变量list_labels
    list_labels = []
    # 使用max_row函数，计算sheet的行数
    sheet_rows = sheet.max_row
    # 通过for循环，读取A列2-sheet_rows行数据
    for cell in sheet["A"][2:sheet_rows]:
        # .value属性获取单元格值，并用append()函数添加进list_labels
        list_labels.append(cell.value)
    # 通过for循环，读取B列2-sheet_rows行数据
    for cell in sheet["B"][2:sheet_rows]:
        # .value属性获取单元格值，并用append()函数添加进list_labels
        list_labels.append(cell.value)

    # 使用set()函数对list_labels去重
    # 使用list()函数对去重后的结果转化为列表
    # 赋值给变量list_nodes
    list_nodes = list(set(list_labels))

    # 定义节点列表，赋值给变量nodes
    nodes = []
    # 通过for循环遍历list_labels
    for i in list_nodes:
        # 定义字典，赋值给变量dic
        dic = {}
        # 为dic创建（"name":"标签名"）键：值对
        dic["name"] = i
        # 使用append()函数将dic添加进nodes
        nodes.append(dic)

    # 定义信息流列表，赋值给变量links
    links = []
    # 通过for循环，读取sheet的2-sheet_rows行数据
    for cell in sheet[2:sheet_rows]:
        # 定义字典，赋值给变量dic
        dic = {}
        # .value属性获取元组索引为0的单元格值，为dic创建（"source":"标签名1"）键：值对
        dic["source"] = cell[0].value
        # .value属性获取元组索引为1的单元格值，为dic创建（"target":"标签名"）键：值对
        dic["target"] = cell[1].value
        # .value属性获取元组索引为2的单元格值，为dic创建（"value":"值"）键：值对
        dic["value"] = cell[2].value
        # 使用append()函数将dic添加进links
        links.append(dic)
    # 使用return返回(nodes,links)
    return (nodes, links)


# 定义函数page_Sankey传入参数nodes_links,title_name
def page_Sankey(nodes_links, title_name):
    # 使用Sankey()函数创建对象赋值给sankey
    # 使用InitOpts()，传入参数theme="dark",bg_color="#253441"，赋值给init_opts
    sankey = Sankey(init_opts=opts.InitOpts(theme="dark", bg_color="#253441"))

    # 为创建的实例增加名字（sankey）、传入实验组节点和信息流列表
    sankey.add("sankey",
               nodes=nodes_links[0],
               links=nodes_links[1],
               # 使用LineStyleOpts()，传入参数opacity=0.3, curve=0.5, color="source"，赋值给linestyle_opt
               linestyle_opt=opts.LineStyleOpts(opacity=0.3, curve=0.5, color="source"),
               # 使用LabelOpts()，传入参数position="right",color="white",font_size=10，赋值给label_opts
               label_opts=opts.LabelOpts(position="right", color="white", font_size=10),
               # 使用SankeyLevelsOpts()，传入参数depth为0-5、itemstyle_opts，赋值给列表levels
               levels=[
                   opts.SankeyLevelsOpts(depth=0,
                                         # 使用ItemStyleOpts()，传入参数color="#FF8947",border_color="#FF8947"
                                         itemstyle_opts=opts.ItemStyleOpts(color="#FF8947", border_color="#FF8947"))
                   , opts.SankeyLevelsOpts(depth=1,
                                           # 使用ItemStyleOpts()，传入参数color="#96D15C",border_color="#96D15C"
                                           itemstyle_opts=opts.ItemStyleOpts(color="#96D15C", border_color="#96D15C"))
                   , opts.SankeyLevelsOpts(depth=2,
                                           # 使用ItemStyleOpts()，传入参数color="#479BED",border_color="#479BED"
                                           itemstyle_opts=opts.ItemStyleOpts(color="#479BED", border_color="#479BED"))
                   , opts.SankeyLevelsOpts(depth=3,
                                           # 使用ItemStyleOpts()，传入参数color="#55C4CA",border_color="#55C4CA"
                                           itemstyle_opts=opts.ItemStyleOpts(color="#55C4CA", border_color="#55C4CA"))
                   , opts.SankeyLevelsOpts(depth=4,
                                           # 使用ItemStyleOpts()，传入参数color="#E7BF4F",border_color="#E7BF4F"
                                           itemstyle_opts=opts.ItemStyleOpts(color="#E7BF4F", border_color="#E7BF4F"))
               ]
               )
    # 使用TitleOpts()，传入参数title，赋值给title_opts
    # 使用LegendOpts()，传入参数is_show=False，赋值给legend_opts
    # 调用set_global_opts()
    sankey.set_global_opts(title_opts=opts.TitleOpts(title=title_name), legend_opts=opts.LegendOpts(is_show=False))
    # 使用return返回sankey
    return sankey


# 创建Page对象，并赋值给page
page = Page(layout=Page.DraggablePageLayout)

# 在add()函数中，分别调用对照组和实验组page_Sankey()函数，图例名分别为"用户路径流转图对照组"和"用户路径流转图实验组"
page.add(page_Sankey(gen_nodesandlinks(sheet_A), "用户路径流转图对照组"), page_Sankey(gen_nodesandlinks(sheet_B), "用户路径流转图实验组"))

# 使用render()生成桑基图页面组合保存到./result/sankey.html
page.render("./result/result_ex0_1_sankey.html")
