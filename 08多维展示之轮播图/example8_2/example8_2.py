import os
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pyecharts.charts import Bar, Timeline
import pyecharts.options as opts
import openpyxl

# 资源路径
resourcesPath = "./resources"
# 月产品线销量报表工作簿
monthSalesBookName = "timeline.xlsx"
# 打开工作簿
monthSalesBook = openpyxl.load_workbook(filename=os.path.join(resourcesPath, monthSalesBookName))  # type: Workbook


def read_sheet(sheet_name: str):
    """
    传入工作表，返回当前工作表中需要用于绘图的数据。

    前置条件：所需的工作簿已打开.

    :param sheet_name: 当前工作表的名称

    :return: 元组(表头-x轴, (商家名1, x轴对应的y轴数据组1), (商家名2, x轴对应的y轴数据组2), ...)

    """

    # 选中工作表
    month_sales_sheet = monthSalesBook[sheet_name]  # type: Worksheet
    # 取出所有原始数据，先不做处理，形如：
    # [(None, '草莓', '芒果', '葡萄', '雪梨', '西瓜', '柠檬', '车厘子'),
    #  ('商家A', 130, 136, 112, 61, 149, 99, 62),
    #  ('商家B', 82, 112, 41, 50, 40, 91, 117)]
    org_data = [org for org in month_sales_sheet.iter_rows(values_only=True)]

    # 提取产品信息列表
    product_list = org_data[0][1:]
    # 提取商家及其销量信息
    sales_info = org_data[1:]

    # 组装
    plot_data = [product_list]
    for info in sales_info:
        plot_data.append(info)

    return tuple(plot_data)


# 绘制轮播图，节点名设置为月份（工作表的名字）
# 柱状图每个数据系名设置为商户的名字：商家X；
# 每个月份图表的标题设置为"月份"+"销量"：X月销量；
# 将图表保存
timeline = Timeline()
# 扫描所有工作表，逐个绘制柱状图加进轮播图里面
for sheetName in monthSalesBook.sheetnames:
    # 取出用于绘图的一组数据：(表头-x轴, (商家名1, x轴对应的y轴数据组1), (商家名2, x轴对应的y轴数据组2), ...)
    packedData = read_sheet(sheetName)

    # 提取x轴数据
    x_data = packedData[0]

    # 提取多组y轴数据字典
    y_data_dict = dict()
    for y_data_item in packedData[1:]:
        # 键值对——'商家名': [对应x轴的系列取值]
        y_data_dict[y_data_item[0]] = y_data_item[1:]

    # 绘柱状图
    bar = (
        Bar()
        .set_global_opts(title_opts=opts.TitleOpts(title=f"{sheetName}销量"))
        .add_xaxis(xaxis_data=x_data)
    )
    for k_vendor, v_y_data in y_data_dict.items():
        bar.add_yaxis(
            series_name=k_vendor,
            y_axis=v_y_data
        )
    # 加进轮播图
    timeline.add(chart=bar, time_point=sheetName)

# 保存轮播图路径
resultPath = "./result"
if not os.path.exists(path=resultPath):
    os.mkdir(path=resultPath)
# 文件名
resultFileName = "monthly_sales_timeline.html"
# 执行保存
timeline.render(path=os.path.join(resultPath, resultFileName))
