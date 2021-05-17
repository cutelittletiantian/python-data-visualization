import os
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pyecharts.charts import Funnel
import pyecharts.options as opts

# 资源文件路径
resourcesPath = "./resources"
salesBookName = "销售漏斗.xlsx"


def read_rows(wb: Workbook=None, sheetname=None) -> list:
    """
    从excel文档中读出所有行，每一行都是列表，所有行存入一个大列表，返回出去
    读取时，不做清洗，只是忠诚地把文件里的数据转移出来到列表里
    :param wb: excel工作簿（已经打开）
    :param sheetname: 工作表Worksheet名
    :return: list
    """
    # 选表
    sheet = wb[sheetname]  # type: Worksheet
    # 读出所有行数据
    row_data = [list(rowContent) for rowContent in sheet.iter_rows(values_only=True)]
    return row_data


# 加载工作簿
salesBook = openpyxl.load_workbook(filename=os.path.join(resourcesPath, salesBookName))  # type: Workbook
# 内有多个sheets，逐个遍历
for teamSheetName in salesBook.sheetnames:
    # 选工作表
    teamSheet = salesBook[teamSheetName]  # type: Worksheet

    # 取出表中数据
    rowData = read_rows(wb=salesBook, sheetname=teamSheetName)
    # 谈判阶段
    stages = rowData[0][1:]
    # 每个阶段达成的数量
    num = rowData[1][1:]

    # 处理数据，[标签项，数值项]处理成形如["目标客户100%", 1000]格式
    # 相对变化率的数值，计算时保留一位小数
    # 计算各个环节人数相对变化率
    pass_rate = [100]
    # 从第2个元素开始，所以下标始于1
    for index in range(1, len(num)):
        relative_percentage = (num[index] / num[index - 1]) * 100
        # 保留一位小数
        relative_percentage = round(relative_percentage, 1)
        pass_rate.append(relative_percentage)

    # 按照格式组装标签项
    label_item = []
    for common_index, rate_item in enumerate(pass_rate):
        label_item.append(f"{stages[common_index]}{rate_item}%")
    # 组装绘图用的数据对组装
    data_pair = []
    for common_index, num_item in enumerate(num):
        data_pair.append(
            [label_item[common_index], num_item]
        )

    # 绘制漏斗图，要求：
    # 1、隐藏图例；
    # 2、将标题设置为：f"{小组名称}销售漏斗"格式
    # 3、将三个图表分别保存到result目录下，命名格式："{小组名称}.html"
    # 4、漏斗项之间的间距为10
    resultPath = "./result"
    if not os.path.exists(resultPath):
        os.mkdir(path=resultPath)
    resultFileName = f"{teamSheetName}_sales_funnel.html"
    funnel = (
        Funnel()
        .set_global_opts(
            title_opts=opts.TitleOpts(title=f"{teamSheetName}销售漏斗"),
            legend_opts=opts.LegendOpts(is_show=False)
        )
        .add(
            series_name="",
            data_pair=data_pair,
            gap=10
        )
        .render(path=os.path.join(resultPath, resultFileName))
    )
