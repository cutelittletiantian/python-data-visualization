import os
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pyecharts.charts import Funnel, Tab
import pyecharts.options as opts

# 资源目录
resourcesPath = "./resources"
# 资源文件
salesBookName = "销售.xlsx"

# 打开excel文件表
salesBook = openpyxl.load_workbook(filename=os.path.join(resourcesPath, salesBookName))  # type: Workbook


def read_data_pair(sheet: Worksheet) -> list:
    """
    从一个工作表中，读出生成一个漏斗图所需的必备数据对。
    :param sheet:正在处理的工作表
    :return:二元列表[标签项，数据项]为[f"{环节}{相对变化率}%", 对应数量]的元素所构成的二维列表。相对变化率保留1位小数
    """
    # 忠诚读取各行数据
    row_data = [list(rowContent) for rowContent in sheet.iter_rows(values_only=True)]
    # 原始数据：环节
    stages = row_data[0][1:]
    # 原始数据：各阶段达成数量
    num = [eval(str(data_item)) for data_item in row_data[1][1:]]

    # 计算各个环节人数相对变化率
    pass_rate = [100]
    # 从第2个元素开始，所以下标始于1
    for index in range(1, len(num)):
        relative_percentage = (num[index] / num[index - 1]) * 100
        # 保留一位小数
        relative_percentage = round(relative_percentage, 1)
        pass_rate.append(relative_percentage)

    # 按照格式组装标签项
    label_item = []
    for common_index, rate_item in enumerate(pass_rate):
        label_item.append(f"{stages[common_index]}{rate_item}%")
    # 组装绘图用的数据对组装
    data_pair = []
    for common_index, num_item in enumerate(num):
        data_pair.append(
            [label_item[common_index], num_item]
        )

    return data_pair


# 用于添加所有绘制出来的漏斗图
tab = Tab()

for monthSheetName in salesBook.sheetnames:
    # 获取该表下用于绘图的数据对
    data_pair = read_data_pair(salesBook[monthSheetName])
    # 绘制漏斗图，要求：
    # 1、隐藏图例；
    # 2、每张图标题设置为：f"{月份}销售"
    # 3、图中每个漏斗项中间的距离为10
    # 4、多漏斗图用看板展示；
    # 5、保存
    funnel = (
        Funnel()
        .set_global_opts(
            title_opts=opts.TitleOpts(title=f"{monthSheetName}销售"),
            legend_opts=opts.LegendOpts(is_show=False)
        )
        .add(
            series_name="",
            data_pair=data_pair,
            gap=10
        )
    )
    tab.add(chart=funnel, tab_name=f"{monthSheetName}")

# 保存看板图
resultPath = "./result"
if not os.path.exists(resultPath):
    os.mkdir(path=resultPath)
resultFileName = "year_funnel_tab.html"

tab.render(path=os.path.join(resultPath, resultFileName))
