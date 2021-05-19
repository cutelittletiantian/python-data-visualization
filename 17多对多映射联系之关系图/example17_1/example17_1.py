import os
from pyecharts.charts import Graph
import pyecharts.options as opts
from pyecharts.commons.utils import JsCode
import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# 使用关系图，可视化Python知识点关联图谱

# 资源文件夹
resourcesPath = "./resources"
# 知识点文件
knowledgeBookName = "知识点.xlsx"
# 存储边的父子节点所在工作表
linkSheetName = "知识点的关系"
# 存储各个知识点带权节点的(节点名, 权重)关系
nodeSheetName = "数量统计"

# 打开文件
knowledgeBook = openpyxl.load_workbook(filename=os.path.join(resourcesPath, knowledgeBookName))  # type: Workbook
# 边工作表
linkSheet = knowledgeBook[linkSheetName]  # type: Worksheet
# 节点工作表
nodeSheet = knowledgeBook[nodeSheetName]  # type: Worksheet

# 读取表格清洗后的数据（通常这个表格数据是清洗后的，得到这个表格前可能需要清洗自然语言）
# (知识点, 知识点)组成的边
orgLinkList = [rowItem for rowItem in linkSheet.iter_rows(min_row=2, values_only=True)]
# (知识点, 权重)组成的节点
orgNodeList = [rowItem for rowItem in nodeSheet.iter_rows(min_row=2, values_only=True)]
print(orgLinkList)
print(orgNodeList)

# 按[{"source": xxx, "target": xxx}, ...]格式组合边的数据
# 原始数据转化为指定格式方法
linkList = [opts.GraphLink(source=linkItem[0], target=linkItem[1]) for linkItem in orgLinkList]
# 按[{"id":"Python", "name":"Python ", "symbolSize":24}]格式组合节点数据，且假设id就是节点名（这里没有重名）
# 注意：这里节点的id是自定义添加的字段，用作唯一标识
# nodeList = [opts.GraphNode(id=nodeItem[0], name=nodeItem[0], symbol_size=nodeItem[1]) for nodeItem in orgNodeList]
nodeList = [{"id": nodeItem[0], "name": nodeItem[0], "symbolSize": nodeItem[1]} for nodeItem in orgNodeList]

# 绘制关系图需求：
# 1. 设置图表的图例系列名称为空；
# 2. 整体呈圆形布局，数据标签可依照节点实际情况进行旋转调整
# 3. 用给定的js代码设置边为渐变色（怎么设置可以查文档）
# 4. 边的弯曲度设为0.3，宽度设为1.5；
# 5. 节点的颜色"#474747"；
# 6. 标题设置为：Python知识图谱；
# 7. 保存图表
# 保存路径
resultPath = "./result"
if not os.path.exists(path=resultPath):
    os.mkdir(path=resultPath)
# 保存文件名
resultFileName = "pythonKG_graph.html"

# 用于控制边渐变颜色的js代码（具体参数含义官方文档中有给出说明）
line_color_js = """
new echarts.graphic.LinearGradient(
    0,0,0,1,
    [{offset: 0,
      color: '#2e9afe'},
     {
      offset: 0.3,
      color: '#01a9db'},
     {
      offset: 0.6,
      color: '#f5bca9'},
     {
      offset: 1,
      color: '#b40404'}]
    )
"""

# 执行绘图
knowledgeGraph = (
    Graph()
    .set_global_opts(title_opts=opts.TitleOpts(title="Python知识图谱"))
    .add(
        series_name="",
        links=linkList,
        nodes=nodeList,
        layout="circular",
        is_rotate_label=True,
        linestyle_opts=opts.LineStyleOpts(
            color=JsCode(js_code=line_color_js),
            curve=0.3,
            width=1.5
        ),
        itemstyle_opts=opts.ItemStyleOpts(color="#474747")
    )
    .render(path=os.path.join(resultPath, resultFileName))
)
